# -*- coding: utf-8 -*-
"""Create review and bibliometric writing templates for sport geography projects."""

from pathlib import Path
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_sheet(ws):
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 46)


def add_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def build_workbook(out_path):
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("PRISMA筛选记录")
    add_rows(
        ws,
        ["阶段", "数量", "排除原因", "备注"],
        [
            ["数据库初检", "", "", "分别记录 WoS、Scopus、PubMed、SPORTDiscus 等"],
            ["过滤后", "", "年份/语言/文献类型/索引", ""],
            ["去重后", "", "DOI/题名重复", ""],
            ["题名摘要筛选后", "", "主题不符", ""],
            ["全文筛选后", "", "无空间/公平/健康或体育地理相关性", ""],
            ["最终纳入", "", "", ""],
        ],
    )

    ws = wb.create_sheet("文献批判性编码")
    add_rows(
        ws,
        [
            "ID",
            "作者年份",
            "题名",
            "期刊",
            "DOI",
            "对象类型",
            "暴露/可达性指标",
            "空间尺度",
            "数据来源",
            "方法家族",
            "人群",
            "正义维度",
            "健康/活动路径",
            "规划政策启示",
            "局限",
            "证据角色",
        ],
        [],
    )

    ws = wb.create_sheet("Claim-Evidence Map")
    add_rows(
        ws,
        [
            "核心主张",
            "证据来源",
            "支持图表",
            "解释逻辑",
            "边界条件",
            "可写入章节",
            "风险/需核验",
        ],
        [],
    )

    ws = wb.create_sheet("期刊定位")
    add_rows(
        ws,
        ["目标期刊", "适配主题", "必须强调", "避免", "备选题目"],
        [
            ["Sustainable Cities and Society", "健康城市/韧性/热暴露/决策支持", "方法链条和政策议程", "纯工具图谱", ""],
            ["Cities", "城市治理/空间正义/公共服务配置", "规划机制和治理意义", "无规划应用的技术分析", ""],
            ["Landscape and Urban Planning", "绿地质量/景观服务/公园公平", "景观规划和设计启示", "过泛健康口号", ""],
            ["Environment International", "暴露-健康路径/环境正义", "证据质量和健康风险", "未经证据支持的因果宣称", ""],
        ],
    )

    ws = wb.create_sheet("图表规划")
    add_rows(
        ws,
        ["图表编号", "要证明的主张", "数据来源", "图表类型", "是否主文", "备注"],
        [
            ["Fig.1", "检索过程和语料范围可靠", "PRISMA记录", "流程图+语料概况", "是", ""],
            ["Fig.2", "领域从身体活动/建成环境转向公平与暴露", "关键词/主题演化", "主题演化/overlay", "是", ""],
            ["Fig.3", "方法从距离可达性走向质量/动态/热暴露", "方法编码", "堆叠图/热图", "是", ""],
            ["Fig.4", "现有研究以分配公平为主，程序和承认正义不足", "正义维度编码", "矩阵/桑基图", "是", ""],
            ["Fig.5", "未来议程需要对象-指标-人群-治理联动", "综合框架", "概念框架图", "是", ""],
        ],
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("output", nargs="?", default="review_bibliometric_templates.xlsx")
    args = parser.parse_args()
    out_path = Path(args.output).resolve()
    build_workbook(out_path)
    print(out_path)


if __name__ == "__main__":
    main()


